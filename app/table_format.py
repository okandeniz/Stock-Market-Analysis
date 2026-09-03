"""Consistent Turkish number/date formatting for HTML result tables."""
from __future__ import annotations

from numbers import Number
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .plotly_theme import format_tr_number


_TWO_DECIMAL_HINTS = (
    "%",
    "oran",
    "marj",
    "devir",
    "süresi",
    "suresi",
    "gün",
    "gun",
    "döngü",
    "dongu",
    "f/k",
    "fd/",
    "pd/",
    "fiyat",
    "hbk",
    "getiri",
    "çarpan",
    "carpan",
    "median",
    "net_borc/favok",
)
_INTEGER_HINTS = ("piotroski", "skor")
_INTEGER_EXACT_NAMES = {
    "hedef yıl",
    "açıklanan çeyrek",
    "tahmin edilen çeyrek",
    "güven puanı",
    "örnek sayısı",
    "cv pencere sayısı",
    "cv ufku",
    "toplam aday",
    "başarılı aday",
    "sektör örnek sayısı",
    "backtest örnek sayısı",
}


def _column_decimals(column: Any, series: pd.Series) -> int:
    name = str(column).casefold()
    if name in _INTEGER_EXACT_NAMES:
        return 0
    if any(hint in name for hint in _INTEGER_HINTS):
        return 0
    if any(hint in name for hint in _TWO_DECIMAL_HINTS):
        return 2

    numeric = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if numeric.empty:
        return 2
    # Büyük TL tutarlarında kuruş gösterme; küçük değerlerde iki ondalık koru.
    return 0 if float(numeric.abs().median()) >= 100_000 else 2


def _format_value(value: Any, decimals: int) -> str:
    if value is None:
        return "—"
    try:
        if bool(pd.isna(value)):
            return "—"
    except (TypeError, ValueError):
        pass
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(value).strftime("%Y-%m")
    if isinstance(value, Number) and not isinstance(value, (bool, np.bool_)):
        return format_tr_number(value, decimals)
    return str(value)


def format_dataframe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Return a string-only display copy; calculations remain untouched."""
    formatted = df.copy()
    for column in formatted.columns:
        if str(column).casefold() == "hedef yıl":
            formatted[column] = formatted[column].map(
                lambda value: "—" if pd.isna(value) else str(int(value))
            )
            continue
        decimals = _column_decimals(column, formatted[column])
        formatted[column] = formatted[column].map(lambda value, d=decimals: _format_value(value, d))

    if isinstance(formatted.index, pd.DatetimeIndex):
        formatted.index = formatted.index.strftime("%Y-%m")
    return formatted


def dataframe_to_html(
    df: pd.DataFrame,
    *,
    index: bool = True,
    classes: str = "data-table",
) -> str:
    """Render a DataFrame with one header row and Turkish display formats.

    ``DataFrame.to_html`` places a named index (for example ``Kod`` or
    ``Kalem``) in a second ``thead`` row. Separate sticky rows overlap while
    scrolling, so HTML tables expose index levels as ordinary first columns.
    """
    formatted = format_dataframe_for_display(df)
    html_index = index
    if index:
        index_level_count = formatted.index.nlevels
        original_names = list(formatted.index.names)
        reserved = {str(column) for column in formatted.columns}
        temporary_names: list[str] = []
        for level in range(index_level_count):
            candidate = f"__html_index_{level}__"
            while candidate in reserved:
                candidate = f"_{candidate}"
            temporary_names.append(candidate)
            reserved.add(candidate)

        flattened = formatted.copy()
        flattened.index = flattened.index.set_names(temporary_names)
        formatted = flattened.reset_index()
        columns = list(formatted.columns)
        for level, original_name in enumerate(original_names):
            if original_name is not None:
                columns[level] = str(original_name)
            elif index_level_count == 1 and isinstance(df.index, pd.DatetimeIndex):
                columns[level] = "Dönem"
            else:
                columns[level] = ""
        formatted.columns = columns
        html_index = False

    return formatted.to_html(
        index=html_index,
        escape=True,
        border=0,
        classes=classes,
        na_rep="—",
    )


def write_formatted_excel(df: pd.DataFrame, path: Path, *, index: bool = True) -> None:
    """Write numeric data with Excel-native formats, filters and frozen headers."""
    from openpyxl.styles import Font, PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=index, sheet_name="Analiz")
        worksheet = writer.sheets["Analiz"]
        worksheet.freeze_panes = "B2" if index else "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="0D6628")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        column_offset = 2 if index else 1
        if index and isinstance(df.index, pd.DatetimeIndex):
            for cell in worksheet.iter_cols(min_col=1, max_col=1, min_row=2):
                for item in cell:
                    item.number_format = "yyyy-mm"

        for position, column in enumerate(df.columns, start=column_offset):
            decimals = _column_decimals(column, df[column])
            number_format = (
                "0"
                if str(column).casefold() == "hedef yıl"
                else "#,##0" if decimals == 0 else "#,##0.00"
            )
            for cell_tuple in worksheet.iter_cols(
                min_col=position,
                max_col=position,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for cell in cell_tuple:
                    if isinstance(cell.value, Number) and not isinstance(cell.value, bool):
                        cell.number_format = number_format

        headers = ([df.index.name or "Dönem"] if index else []) + [str(column) for column in df.columns]
        for position, header in enumerate(headers, start=1):
            sample_values = [str(cell.value or "") for cell in list(worksheet.columns)[position - 1][:100]]
            worksheet.column_dimensions[worksheet.cell(1, position).column_letter].width = min(
                max(len(header), *(len(value) for value in sample_values)) + 2,
                36,
            )
